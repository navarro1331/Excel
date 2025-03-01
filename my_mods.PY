﻿import pandas as pd
from fuzzywuzzy import fuzz, process
import pyautogui
import time
import openpyxl


class CityIdFinder:
    def __init__(self, city_id_file_path, city_id_sheet_name='City Id Hub'):
        self.city_id_file_path = city_id_file_path
        self.city_id_sheet_name = city_id_sheet_name
        self.city_id_df = None

    def load_city_id_data(self):
        try:
            self.city_id_df = pd.read_excel(self.city_id_file_path, sheet_name=self.city_id_sheet_name)
            print(f"'{self.city_id_sheet_name}' sheet loaded successfully from '{self.city_id_file_path}'.")
        except Exception as e:
            print(f"Error loading City Id data: {e}")
            self.city_id_df = None

    def preprocess_search_phases(self, search_value):
        """Returns the search phases for exact matching."""
        base_value = search_value.split('(')[0].rstrip('.').strip()  # Base for search phases
        return [
            search_value,  # Phase 1: Original value
            base_value,  # Phase 2: Remove text after '(' and strip
            base_value.replace(',', ''),  # Phase 3: Remove commas
            base_value.replace('.', ''),  # Phase 4: Remove periods
            base_value.split(',')[0].strip()  # Phase 5: Remove text after the first comma
        ]

    def find_city_id(self, contractor_name, threshold=75):
        """
        Finds the City ID using exact and fuzzy matching.
        :param contractor_name: The name to search for
        :param threshold: Fuzzy match score threshold (default: 75)
        :return: City ID if found, otherwise "Not Found"
        """
        if self.city_id_df is None:
            return "Data not loaded"

        columns_to_search = self.city_id_df.columns[1:]  # Adjust based on the actual columns you want to search

        # Preprocess the search phases
        search_phases = self.preprocess_search_phases(contractor_name)

        # Try exact matching with different search phases
        for phase, trimmed_value in enumerate(search_phases, 1):
            print(f"Phase {phase} - Searching for: {trimmed_value}")
            for col in columns_to_search:
                # Convert the column to string to avoid errors when using .str
                column_data = self.city_id_df[col].astype(str)

                matches = column_data.str.strip().eq(trimmed_value)
                if matches.any():  # Use vectorized matching for efficiency
                    match_index = matches.idxmax()
                    print(f"Found exact match in column {col} at phase {phase}, index {match_index}")
                    return self.city_id_df.iloc[match_index, 0]  # Return City ID

        # If no exact match is found, attempt fuzzy matching
        print("No exact match found, attempting fuzzy match...")
        best_match, best_score = None, 0

        for col in columns_to_search:
            try:
                # Convert the column to string to handle any non-string values
                possible_values = self.city_id_df[col].dropna().astype(str).tolist()
                match, score = process.extractOne(contractor_name, possible_values, scorer=fuzz.token_sort_ratio)
                if score > best_score:
                    best_score = score
                    best_match = match
                    match_index = self.city_id_df[col].eq(best_match).idxmax()
                    best_city_id = self.city_id_df.iloc[match_index, 0]

            except Exception as e:
                print(f"Error during fuzzy matching: {e}")
                continue

        if best_score >= threshold:
            print(f"Fuzzy match found: {best_match} with score {best_score}")
            return best_city_id

        print("No match found.")
        return "Not Found"

def find_value_in_column(df, search_column, search_values, return_column):
    """
    Searches for matching values in a specified column and returns corresponding values from another column.
    If a value is not found, returns 'Not Found' for that value.
    """
    results = []

    for value in search_values:
        try:
            # Use df.loc for exact matching instead of idxmax
            matched_rows = df.loc[df[search_column] == value]
            
            if not matched_rows.empty:
                # If there are multiple matches, return the first match
                return_value = matched_rows[return_column].iloc[0]
            else:
                # No match found for the current value
                return_value = "Not Found"
        except (ValueError, IndexError, KeyError) as e:
            # Handle any possible exception and return "Not Found"
            return_value = "Not Found"
        
        # Append the search value and the found (or not found) result
        results.append([value, return_value])

    # Return results as a DataFrame
    return pd.DataFrame(results, columns=[search_column, return_column])

def extract_unique_city_ids(file_path, sheet_name, city_id_column):
    """
    Extracts unique City IDs from the specified sheet in the Excel file.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        unique_city_ids = df[city_id_column].dropna().unique()
        return pd.DataFrame(unique_city_ids, columns=[city_id_column])
    except Exception as e:
        print(f"Error in extracting unique city IDs: {e}")
        return None

def merge_city_dataframes(df1, df2, city_id_column, merge_columns):
    """
    Merges two DataFrames on the City Id column and removes duplicates.
    """
    try:
        merged_df = pd.merge(df1, df2[[city_id_column] + merge_columns], on=city_id_column, how='inner')
        return merged_df.drop_duplicates()
    except Exception as e:
        print(f"Error in merging DataFrames: {e}")
        return None

def sum_contract_amounts(merged_df, review_df, city_id_column, contract_column):
    """
    Sums the contract amounts for each City Id from the review DataFrame.
    """
    try:
        pop_merged_df = pd.merge(merged_df, review_df[[city_id_column, contract_column]], on=city_id_column, how='inner')
        return pop_merged_df.groupby(city_id_column).agg({contract_column: 'sum'}).reset_index()
    except Exception as e:
        print(f"Error in summing contract amounts: {e}")
        return None

def add_pop_eligible_column(df, contract_column, deviated_as_column):
    try:
        df['POP Eligible'] = df.apply(
            lambda row: 'Yes' if row[contract_column] > 200000 and row[deviated_as_column] != 'Supplier' else 'No',
            axis=1
        )
        return df
    except Exception as e:
        print(f"Error in adding POP Eligible column: {e}")
        return df

def append_found_information(hub_df, fvic_df, city_id_column):
    """
    Appends found information to the existing DataFrame, avoiding duplicate 'City Id' columns.
    """
    try:
        fvic_df = fvic_df.drop(columns=[city_id_column])  # Drop the duplicate City Id column
        return pd.concat([hub_df, fvic_df], axis=1)
    except Exception as e:
        print(f"Error in appending found information: {e}")
        return hub_df

def write_to_excel(df, file_path, sheet_name, mode='a'):
    """
    Writes a DataFrame to an Excel file.
    """
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        print(f"Data written to {sheet_name} in {file_path}")
    except Exception as e:
        print(f"Error in writing to Excel: {e}")

def ghost_ci(
    input_file_path_1,           # Path to the first input Excel file (e.g., contractor data file)
    city_id_file_path,           # Path to city ID data file
    input_sheet_1,               # Name of the first sheet (contractor or data to be processed)
    output_sheet_2,              # Name of the second sheet (sheet where merged data will be saved or merged)
    ci_hub,                      # Name of the sheet containing City ID information
    lookup_column,               # Name of the column in sheet_1 to look up values from the lookup table (e.g., contractor column)
    ci_column='ID',              # Name of the column containing city IDs in both tables
    columns_to_merge=None,       # List of columns to merge from the lookup table (default is None)
    merge_on='City Id',          # Column to perform the merge on, default is the unique identifier column
    lookup_value_column='Value', # Column to merge from sheet_1 (e.g., column with values like numbers or contracts)
    output_column=None,          # Optional: Rename the output column after merging (default is None)
    merge_strategy='left',       # Merge strategy, default is 'left' join
    output_file_path=None,       # Output file path; if not specified, defaults to input_file_path_1
    output_sheet_name=None,      # Name of the output sheet; if not specified, defaults to input_sheet_2
    overwrite_sheet=True         # If True, overwrites the sheet in the output file, otherwise appends
):
    """
    This function processes data by finding and merging values based on a lookup column and city IDs,
    then writes the merged data back to an Excel file.
    
    Args:
        input_file_path_1 (str): Path to the Excel file containing the primary data for merging.
        city_id_file_path (str): Path to city ID data file.
        input_sheet_1 (str): Name of the sheet in the primary data file.
        output_sheet_2 (str): Name of the sheet where merged data will be saved.
        ci_hub (str): Name of the sheet containing City ID information in the city ID data file.
        lookup_column (str): Column name in input_sheet_1 to look up values from the city ID lookup table.
        ci_column (str): Column containing city IDs in both datasets.
        columns_to_merge (list): Columns to merge from the lookup table (default is None).
        merge_on (str): Column to merge on (default is 'City Id').
        lookup_value_column (str): Name of the column in the lookup data to merge.
        output_column (str): If specified, the output column will be renamed (default is None).
        merge_strategy (str): Merge strategy ('left', 'right', 'inner', 'outer') (default is 'left').
        output_file_path (str): Output file path. If None, it defaults to input_file_path_1.
        output_sheet_name (str): Output sheet name. If None, it defaults to output_sheet_2.
        overwrite_sheet (bool): Whether to overwrite the existing sheet. Default is True.
    
    Returns:
        None
    """
    # Step 1: Load data from the primary input file
    try:
        primary_df = pd.read_excel(input_file_path_1, sheet_name=output_sheet_2)
        lookup_df = pd.read_excel(input_file_path_1, sheet_name=input_sheet_1)
    except Exception as e:
        print(f"Error reading Excel files: {e}")
        return

    # Step 2: Load city ID lookup data using CityIdFinder
    lookup_finder = CityIdFinder(city_id_file_path, ci_hub)
    lookup_finder.load_city_id_data()

    # Step 3: Add a new column with city IDs based on the lookup_column values
    try:
        lookup_df[ci_column] = lookup_df[lookup_column].apply(lookup_finder.find_city_id)
    except Exception as e:
        print(f"Error while applying city ID lookup: {e}")
        return

    # Step 4: Verify if the lookup_value_column exists and prepare for merging
    if lookup_value_column in lookup_df.columns:
        if columns_to_merge is None:
            columns_to_merge = [ci_column, lookup_value_column]  # Default columns to merge

        # Merge lookup_df into primary_df based on the specified column
        try:
            merged_df = pd.merge(primary_df, lookup_df[columns_to_merge], on=merge_on, how=merge_strategy)

            # Rename the column if output_column is specified
            if output_column:
                merged_df.rename(columns={lookup_value_column: output_column}, inplace=True)

        except KeyError as ke:
            print(f"KeyError during merging: {ke}")
            return
        except Exception as e:
            print(f"Error during merging: {e}")
            return

        # Step 5: Determine output file path and sheet name
        if not output_file_path:
            output_file_path = input_file_path_1  # Use input file if no output file path is provided

        if not output_sheet_name:
            output_sheet_name = output_sheet_2  # Use output_sheet_2 if no output sheet name is provided

        # Step 6: Write the combined data to the Excel file
        try:
            mode = 'w' if overwrite_sheet else 'a'
            write_to_excel(merged_df, output_file_path, output_sheet_name, mode=mode)
        except Exception as e:
            print(f"Error writing to Excel: {e}")
    else:
        print(f"Error: Column '{lookup_value_column}' not found in lookup_df.")

def track_mouse_position(update_interval=0.1):
    '''
if __name__ == "__main__":
    track_mouse_position(0.2)
'''
    print("Press Ctrl+C to stop.")
    try:
        while True:
            x, y = pyautogui.position()
            print(f"Mouse position: ({x}, {y})", end="\r")
            time.sleep(update_interval)
    except KeyboardInterrupt:
        print("\nStopped tracking.")

def read_data(file_path, sheet_name, start_row=1, start_col=1, end_row=None, end_col=None):
    """
    Read data from an Excel sheet within a specified range.

    Parameters:
    - file_path (str): Path to the Excel file.
    - sheet_name (str): Name of the sheet to read data from.
    - start_row (int): Starting row for reading data (1-indexed).
    - start_col (int): Starting column for reading data (1-indexed).
    - end_row (int): Ending row for reading data (1-indexed, optional).
    - end_col (int): Ending column for reading data (1-indexed, optional).

    Returns:
    - data (list of lists): Extracted data organized as a 2D list.
    """
    try:
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}.")

        worksheet = workbook[sheet_name]
        data = []

        # Define the range of cells to read
        end_row = end_row or worksheet.max_row
        end_col = end_col or worksheet.max_column

        # Extract data within the specified range
        for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            data.append([cell.value for cell in row])

        return data

    except Exception as e:
        print(f"An error occurred while reading data: {e}")
        return []

def paste_data_as_values(file_path, sheet_name, data, start_row=1, start_col=1):
    """
    Paste data into an Excel sheet without changing existing formatting.

    Parameters:
    - file_path (str): Path to the Excel file.
    - sheet_name (str): Name of the sheet to paste data into.
    - data (list of lists): Data to paste, organized as a 2D list.
    - start_row (int): Row number to start pasting data (1-indexed).
    - start_col (int): Column number to start pasting data (1-indexed).
    
    Example:
    >>> paste_data_as_values("example.xlsx", "Sheet1", [["A1", "B1"], ["A2", "B2"]], 2, 2)
    """
    try:
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}.")
        
        worksheet = workbook[sheet_name]
        
        # Paste data as values without changing existing formatting
        for i, row in enumerate(data, start=start_row):
            for j, value in enumerate(row, start=start_col):
                worksheet.cell(row=i, column=j).value = value

        # Save the workbook to preserve changes
        workbook.save(file_path)
        print(f"Data pasted successfully into '{file_path}', sheet '{sheet_name}'.")
    
    except Exception as e:
        print(f"An error occurred: {e}")







